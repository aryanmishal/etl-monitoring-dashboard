import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import io

def create_summary_excel(summary_data, view_type, date_range):
    """
    Create an Excel file with summary data
    
    Args:
        summary_data (dict): Summary data from the API
        view_type (str): 'daily', 'weekly', or 'monthly'
        date_range (str): Date range string for display
    
    Returns:
        bytes: Excel file as bytes
    """
    try:
        print(f"Creating Excel file - View Type: {view_type}, Date Range: {date_range}")
        print(f"Summary data keys: {list(summary_data.keys()) if summary_data else 'None'}")
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{view_type.title()} Summary"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_font = Font(bold=True, color="000000")
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = f"ETL Monitoring - {view_type.title()} Summary Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        # Date range - Convert to DD-MM-YYYY format
        def format_date_range(date_range_str):
            """Convert date range from YYYY-MM-DD to DD-MM-YYYY format"""
            if ' to ' in date_range_str:
                # Handle date ranges like "2025-07-28 to 2025-08-03"
                start_date, end_date = date_range_str.split(' to ')
                start_formatted = datetime.strptime(start_date, '%Y-%m-%d').strftime('%d-%m-%Y')
                end_formatted = datetime.strptime(end_date, '%Y-%m-%d').strftime('%d-%m-%Y')
                return f"{start_formatted} to {end_formatted}"
            else:
                # Handle single date like "2025-07-28"
                return datetime.strptime(date_range_str, '%Y-%m-%d').strftime('%d-%m-%Y')
        
        formatted_date_range = format_date_range(date_range)
        ws['A2'] = f"Period: {formatted_date_range}"
        ws['A2'].font = Font(bold=True, size=12)
        ws.merge_cells('A2:B2')
        
        # Generated timestamp
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(italic=True, size=10)
        ws.merge_cells('A3:B3')
        
        # Add some spacing
        ws['A4'] = ""
        ws['A5'] = ""
        
        # Key Metrics Section
        row = 6
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        metrics_data = [
            ["Metric", "Value"],
            ["Total Raw Records", summary_data.get('total_raw', 0)],
            ["Total Bronze Records", summary_data.get('total_bronze', 0)],
            ["Total Silver Records", summary_data.get('total_silver', 0)]
        ]
        
        for i, row_data in enumerate(metrics_data):
            current_row = row + i
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                if i == 0:  # Header row
                    cell.font = subheader_font
                    cell.fill = subheader_fill
        
        row += len(metrics_data)
        ws[f'A{row}'] = ""
        row += 1
        
        # Ingestion Status Section
        ws[f'A{row}'] = "Ingestion Status"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ingestion_data = [
            ["Metric", "Value"],
            ["Total Users", summary_data.get('total_users', 0)],
            ["Successful Ingestions", summary_data.get('successful_ingestions', 0)],
            ["Missing Ingestions", summary_data.get('total_users', 0) - summary_data.get('successful_ingestions', 0)]
        ]
        
        for i, row_data in enumerate(ingestion_data):
            current_row = row + i
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                if i == 0:  # Header row
                    cell.font = subheader_font
                    cell.fill = subheader_fill
        
        row += len(ingestion_data)
        ws[f'A{row}'] = ""
        row += 1
        
        # Pipeline Status Section
        ws[f'A{row}'] = "Pipeline Status"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        pipeline_data = [
            ["Pipeline Stage", "Status"],
            ["Raw to Bronze", summary_data.get('raw_to_bronze_status', 'Unknown')],
            ["Bronze to Silver", summary_data.get('bronze_to_silver_status', 'Unknown')],
            ["Overall Status", "Success" if (
                summary_data.get('raw_to_bronze_status') == 'Success' and
                summary_data.get('bronze_to_silver_status') == 'Success' and
                summary_data.get('total_users', 0) == summary_data.get('successful_ingestions', 0)
            ) else "Failed"]
        ]
        
        for i, row_data in enumerate(pipeline_data):
            current_row = row + i
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                if i == 0:  # Header row
                    cell.font = subheader_font
                    cell.fill = subheader_fill
                elif col == 2 and i > 0:  # Status column
                    if value == 'Success':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == 'Failed':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row += len(pipeline_data)
        ws[f'A{row}'] = ""
        row += 1
        
        # Users List Section (if available)
        if 'users' in summary_data and summary_data['users']:
            ws[f'A{row}'] = "Users"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:B{row}')
            
            row += 1
            # Header for user IDs
            ws[f'A{row}'] = "User ID"
            ws[f'A{row}'].font = subheader_font
            ws[f'A{row}'].fill = subheader_fill
            ws[f'A{row}'].border = border
            ws.merge_cells(f'A{row}:B{row}')
            
            row += 1
            # Each user ID on its own row, spanning both columns
            for user_id in summary_data['users']:
                ws[f'A{row}'] = user_id
                ws[f'A{row}'].border = border
                ws.merge_cells(f'A{row}:B{row}')
                row += 1
        
        # Set specific column widths for better formatting
        ws.column_dimensions['A'].width = 35  # Metric/Pipeline Stage column
        ws.column_dimensions['B'].width = 25  # Value/Status column
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        result = output.getvalue()
        print(f"Excel file created successfully, size: {len(result)} bytes")
        return result
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        import traceback
        traceback.print_exc()
        raise e 